"""Extract text (and table metadata) from uploaded documents."""

from __future__ import annotations

import csv
import io


def _from_csv(raw: str) -> dict:
    rows = list(csv.reader(io.StringIO(raw)))
    rows = [r for r in rows if any(c.strip() for c in r)]
    header = rows[0] if rows else []
    body = rows[1:]
    table = {
        "linhas": len(body),
        "colunas": len(header),
        "cabecalho": header,
        "amostra": body[:5],
    }
    linhas_txt = [", ".join(header)] + [", ".join(r) for r in body[:50]]
    return {"text": "\n".join(linhas_txt), "table": table}


def _from_pdf(data: bytes) -> str:
    from pypdf import PdfReader  # import tardio (dependência opcional)
    reader = PdfReader(io.BytesIO(data))
    return "\n".join((page.extract_text() or "") for page in reader.pages).strip()


def _from_xlsx(data: bytes) -> dict:
    from openpyxl import load_workbook  # import tardio
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[("" if c is None else str(c)) for c in row]
            for row in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(c.strip() for c in r)]
    header = rows[0] if rows else []
    body = rows[1:]
    table = {"linhas": len(body), "colunas": len(header),
             "cabecalho": header, "amostra": body[:5]}
    linhas_txt = [", ".join(header)] + [", ".join(r) for r in body[:50]]
    return {"text": "\n".join(linhas_txt), "table": table}


def extract(filename: str, data: bytes) -> dict:
    """Return {kind, text, table, warning} from an uploaded file."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else "txt"
    try:
        if ext in {"txt", "md", "log", "text"}:
            return {"kind": "texto", "text": data.decode("utf-8", "replace"),
                    "table": None, "warning": ""}
        if ext == "csv":
            out = _from_csv(data.decode("utf-8", "replace"))
            return {"kind": "csv", "text": out["text"], "table": out["table"], "warning": ""}
        if ext == "pdf":
            return {"kind": "pdf", "text": _from_pdf(data), "table": None, "warning": ""}
        if ext in {"xlsx", "xlsm"}:
            out = _from_xlsx(data)
            return {"kind": "excel", "text": out["text"], "table": out["table"], "warning": ""}
    except ImportError:
        return {"kind": ext, "text": "", "table": None,
                "warning": f"Para ler .{ext} instale as dependências (pypdf/openpyxl)."}
    except Exception as exc:  # noqa: BLE001
        return {"kind": ext, "text": "", "table": None,
                "warning": f"Não consegui ler o arquivo .{ext}: {exc}"}
    # Extensão desconhecida → tenta como texto.
    return {"kind": "texto", "text": data.decode("utf-8", "replace"),
            "table": None, "warning": ""}
